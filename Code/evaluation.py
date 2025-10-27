import os
import numpy as np
import scipy.sparse as sp
import torch
import torch.nn.functional as F
from sklearn import metrics
from openpyxl import load_workbook  # Replace xlrd with openpyxl
import csv

root = os.path.dirname(os.path.dirname(__file__))
path_hpo2id = root + "/data/processed/dis2id.txt"
path_gene2id = root + "/data/processed/entity2id.txt"
path_pos = root + "/data/processed/positive.xlsx"
path_neg1 = root + "/data/processed/random1.xlsx"
path_neg2 = root + "/data/processed/random2.xlsx"

########################################################################
# Disease Similarity Evaluation
########################################################################
#: Loads disease-to-gene mapping
def load_d2g(path):
    return sp.load_npz(path).todense()

# Loads disease ID mappings
def load_dmap(path):
    dmap = dict()
    inv = dict()
    with open(path) as f:
        f.readline()
        for line in f:
            dis, id = line.strip().split()
            dmap[dis] = (int)(id)
            inv[(int)(id)] = dis
    return dmap, inv

def load_disid(path):
    idmap = dict()
    with open(path, "r") as f:
        f.readline()
        for line in f:
            line = line.strip().split("\t")
            disid, name = line[0], line[1]
            if disid not in idmap.keys():
                idmap[disid] = name
            else:
                continue
    return idmap


# Updated load_xlsx function using openpyxl
def load_xlsx(path, dmap):
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = wb.active  # Get the first sheet
    dataset = []
    for row in sheet.iter_rows(min_row=1):  # Iterate through all rows
        dis1 = row[1].value  # Column B (0-based index 1)
        dis2 = row[3].value  # Column D (0-based index 3)
        try:
            d1 = int(dmap[dis1])
            d2 = int(dmap[dis2])
            dataset.append((d1, d2))
        except:
            continue
    return dataset

# Loads evaluation data from csv files
def load_csv(path, dmap):
    dataset = []
    with open(path, 'r') as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            try:
                # Assuming the CSV has exactly two columns
                phen1 = row[0]  # First column
                phen2 = row[1]  # Second column
                
                # Convert values using the mapping
                p1 = int(dmap[phen1])
                p2 = int(dmap[phen2])
                dataset.append((p1, p2))
            except:
                continue
    return dataset

# Calculates disease similarity metrics (AUROC, Average Precision)
def get_disease_auc(dis_embeddings, dis_path, plot=False): 
    """
    Calculate disease similarity AUC metrics
    :param dis_sim: disease embeddings matrix
    :param plot: whether to draw ROC
    :return: AUROC, AP score
    """
    dmap, _ = load_dmap(dis_path+"/dis2id.txt")
    pos_data = load_xlsx(path_pos, dmap)
    neg_data1 = load_xlsx(path_neg1, dmap)
    neg_data2 = load_xlsx(path_neg2, dmap)
    data = pos_data + neg_data1 + neg_data2
    data = np.asarray(data)

    y = np.zeros(len(data), dtype=np.int32)
    for i in range(len(pos_data)):
        y[i] = 1

    x = F.cosine_similarity(dis_embeddings[data[:, 0]], dis_embeddings[data[:, 1]]).view(-1).numpy()

    auroc = metrics.roc_auc_score(y, x)
    ap = metrics.average_precision_score(y, x)

    return auroc, ap

# Function to load and validate with full_dataset.txt
def validate_with_full_dataset(dataset_path, encoder_mode, gamma, dis_embeddings=None, dis_embeddings_1=None, dis_embeddings_2=None):
    """
    Validate model with a full dataset of disease pairs
    
    Args:
        dataset_path: Path to the full dataset file with disease pairs and binary labels
        combined: Flag indicating whether to use combined embeddings
        dis_embeddings: Disease embeddings matrix when using single encoder
        dis_embeddings_1: Disease embeddings from first encoder when combining
        dis_embeddings_2: Disease embeddings from second encoder when combining
        
    Returns:
        auroc: Area under ROC curve
        auprc: Area under Precision-Recall curve
    """
    print(f"Loading validation dataset from {dataset_path}")
    
    # Parse the tab-separated dataset file with disease pair indices and binary labels
    disease_pairs = []
    labels = []
    
    with open(dataset_path, 'r') as f:
        for line in f:
            parts = line.strip().split('\t')
            dis1_idx = int(parts[0])
            dis2_idx = int(parts[1])
            label = int(parts[2])
                    
            disease_pairs.append((dis1_idx, dis2_idx))
            labels.append(label)


    # Convert to numpy arrays
    disease_pairs = np.array(disease_pairs)
    labels = np.array(labels)
    print(f"Loaded {len(disease_pairs)} disease pairs for validation")
    
    # Calculate similarity scores

    if encoder_mode == 'combined':
        print('Using combined embeddings')
      
        # # Average the similarities from both embeddings
        # similarity_scores_1 = F.cosine_similarity(
        #     dis_embeddings_1[disease_pairs[:, 0]], 
        #     dis_embeddings_1[disease_pairs[:, 1]]
        # ).view(-1).numpy()
            
        # similarity_scores_2 = F.cosine_similarity(
        #     dis_embeddings_2[disease_pairs[:, 0]], 
        #     dis_embeddings_2[disease_pairs[:, 1]]
        # ).view(-1).numpy()
            
        #     # Average the two similarity scores
        # similarity_scores = (similarity_scores_1 + similarity_scores_2) / 2.0

        # Average the similarities from both embeddings
        
        combined_emb_1 = torch.cat((gamma * dis_embeddings_1[disease_pairs[:, 0]], (1-gamma) * dis_embeddings_2[disease_pairs[:, 0]]), dim=1)
        combined_emb_2 = torch.cat((gamma * dis_embeddings_1[disease_pairs[:, 1]], (1-gamma) * dis_embeddings_2[disease_pairs[:, 1]]), dim=1)
        # Average the two similarity scores
        similarity_scores = F.cosine_similarity(combined_emb_1, combined_emb_2).view(-1).numpy()
    else:

        similarity_scores = F.cosine_similarity(
            dis_embeddings[disease_pairs[:, 0]], 
            dis_embeddings[disease_pairs[:, 1]]
        ).view(-1).numpy()

    
    # Calculate metrics
    auroc = metrics.roc_auc_score(labels, similarity_scores)
    auprc = metrics.average_precision_score(labels, similarity_scores)
    roc = metrics.roc_curve(labels, similarity_scores)
    precision, recall, _ = metrics.precision_recall_curve(labels, similarity_scores)
    return auroc, auprc, roc, precision, recall
